import os
import time
import pythoncom
import win32com.client as win32
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
WATCH_FOLDER = os.path.join(BASE_DIR, "input")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "output")

def fix_arabic(text):
    try:
        return text.encode("latin-1").decode("windows-1256")
    except:
        return text

def decode_arabic_in_xlsx(xlsx_path):
    try:
        wb = load_workbook(xlsx_path)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = fix_arabic(cell.value)
        wb.save(xlsx_path)
        print(f"✅ Arabic text fixed in: {xlsx_path}")
    except Exception as e:
        print(f"❌ Error fixing Arabic text in {xlsx_path}: {e}")

def convert_excel_to_pdf(xlsx_path):
    pythoncom.CoInitialize()
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = excel.Workbooks.Open(os.path.abspath(xlsx_path))

        relative_path = os.path.relpath(xlsx_path, WATCH_FOLDER)
        folder_part = os.path.dirname(relative_path)
        output_folder = os.path.join(OUTPUT_FOLDER, folder_part)
        os.makedirs(output_folder, exist_ok=True)

        pdf_name = os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf"
        pdf_path = os.path.join(output_folder, pdf_name)

        workbook.ExportAsFixedFormat(0, os.path.abspath(pdf_path))

        workbook.Close(False)
        excel.Quit()

        print(f"✅ Converted to PDF: {pdf_path}")
    except Exception as e:
        print(f"❌ Error converting {xlsx_path} to PDF: {e}")
    finally:
        try:
            del workbook
            del excel
        except:
            pass
        pythoncom.CoUninitialize()

class Handler(FileSystemEventHandler):
    def process(self, event):
        path = event.src_path
        if event.is_directory or "~$" in path or not path.lower().endswith(".xlsx"):
            return
        print(f"📂 Detected: {path}")
        time.sleep(2)
        decode_arabic_in_xlsx(path)
        convert_excel_to_pdf(path)

    def on_created(self, event):
        self.process(event)

    def on_modified(self, event):
        self.process(event)

def convert_existing_files():
    print("📋 Checking existing subfolders for XLSX files...")
    for root, dirs, files in os.walk(WATCH_FOLDER):
        for f in files:
            if f.lower().endswith(".xlsx") and "~$" not in f:
                full_path = os.path.join(root, f)
                print(f"🔄 Converting existing: {full_path}")
                decode_arabic_in_xlsx(full_path)
                convert_excel_to_pdf(full_path)

def watch_folder():
    os.makedirs(WATCH_FOLDER, exist_ok=True)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    convert_existing_files()
    obs = Observer()
    obs.schedule(Handler(), WATCH_FOLDER, recursive=True)
    obs.start()
    print(f"👀 Watching folder & subfolders: {WATCH_FOLDER}")
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        obs.stop()
    obs.join()

if __name__ == "__main__":
    watch_folder()